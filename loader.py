import openpyxl, json
from collections import OrderedDict
from hashlib import sha256
from pprint import pprint

all_classes = ['Accursed', 'Æthera', 'Astromancer', 'Bard', 'Cleric', 'Druid', 'Inquisitor', 'Occultist', 'Odic', 'Odysseer', 'Paladin', 'Ranger', 'Runeshaper', 'Shaman', 'Sorcerer', 'Warden', 'Warlock', 'Wizard']
default_wb = "Spells.xlsx"

def read_row(ws, row):
    spell = OrderedDict()
    for cell in ws.iter_cols(min_row=row, max_row=row):
        cell = cell[0]
        header = ws[cell.column+"1"].value
        spell[header] = cell.value
    return spell

class Spell:
    def __init__(self):
        pass

    @classmethod
    def from_row(cls, row):
        spell = cls()
        spell.name = row['Spell']
        assert type(spell.name) == str and len(spell.name) > 0
        # Fancy dict comprehension. Reads through all classes, True if an x is listed, False otherwise
        spell.classes = {cls: row.get(cls, False) == "x" for cls in all_classes}
        spell.level = row['Level']
        assert type(spell.level) == int
        spell.origin = row['Origin']
        spell.school = row['Sch']
        spell.ritual = row['Ritual'] == "Yes"
        spell.time = row['Time']
        spell.range = row['Range']

        spell.compstr = row['Comp']
        spell.components = {
            "verbal": True if "V" in spell.compstr else None,
            "somantic": True if "S" in spell.compstr else None,
            "material": row['Components'] if "M" in spell.compstr else None
        }

        spell.duration = row['Duration']
        spell.description = row['Full Description/Flavour Text']
        return spell

    @classmethod
    def from_dict(cls, data):
        spell = cls()
        for key in data:
            spell.__setattr__(key, data[key])
        return spell

    def __eq__(self, other):
        return self.__dict__ == other.__dict__

    def __repr__(self):
        return self.name

    def __hash__(self): # https://stackoverflow.com/questions/5884066/hashing-a-dictionary
        return int(sha256(repr(sorted(self.__dict__.items())).encode("utf-8")).hexdigest(), 16)

class Spellbook:

    def __init__(self):
        self.spells = []

    @classmethod
    def from_list(cls, spells):
        spellbook = cls()
        spellbook.spells = spells
        return spellbook

    @classmethod
    def from_json(cls, data):
        spellbook = cls()
        spells = []
        data = json.loads(data)
        for spell in data:
            spell = Spell.from_dict(spell)
            spells.append(spell)
        spellbook.spells = spells
        return spellbook

    @classmethod
    def from_cache(cls, filename):
        with open(filename) as f:
            data = f.read()
        return cls.from_json(data)

    @classmethod
    def from_workbook(cls, filename):
        spellbook = cls()
        wb = openpyxl.load_workbook(filename=filename)
        ws = None
        for worksheet in wb.worksheets:
            if worksheet.title == "Spells":
                ws = worksheet
        assert ws != None
        spells = []
        for row in range(1, ws.max_row):
            rowdata = read_row(ws, row)
            try:
                spell = Spell.from_row(rowdata)
                spells.append(spell)
            except:
                pass
        spellbook.spells = spells
        return spellbook

    def to_json(self):
        return json.dumps([spell.__dict__ for spell in self.spells])

    def to_cache(self, filename):
        data = self.to_json()
        with open(filename, "w") as f:
            f.write(data)

    def search(self, condition):
        return [x for x in self.spells if condition(x)]

    def search_class(self, cls):
        assert cls in all_classes
        return [x for x in self.spells if x.classes[cls]]

    def __eq__(self, other):
        return self.spells == other.spells

def main():
    pass

if __name__ == "__main__": main()